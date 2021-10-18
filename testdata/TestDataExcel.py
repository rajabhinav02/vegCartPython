import openpyxl

class TestData:
    @staticmethod
    def testdataa(tcname):
        excelbook = openpyxl.load_workbook("C:\\Users\\Punam\\workspace_python\\vegcart\\testdata\\vegcart.xlsx")
        sheet = excelbook.active
        Dict ={}

        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column= 1).value == tcname:
                for j in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1, column =j).value]= sheet.cell(row=i, column=j).value
        return [Dict]





